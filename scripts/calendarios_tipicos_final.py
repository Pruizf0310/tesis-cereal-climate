from __future__ import annotations

import math
import re
import unicodedata
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\paola\Tesis\03_Resultados\Fenologia\geoglam_cm4ew_tabla_maestra.xlsx")
ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
OUT.mkdir(exist_ok=True)

MONTHS = [
    ("enero", 1, 31), ("febrero", 32, 59), ("marzo", 60, 90),
    ("abril", 91, 120), ("mayo", 121, 151), ("junio", 152, 181),
    ("julio", 182, 212), ("agosto", 213, 243), ("septiembre", 244, 273),
    ("octubre", 274, 304), ("noviembre", 305, 334), ("diciembre", 335, 366),
]
MONTH_NAMES = [m[0] for m in MONTHS]

CROP_TARGETS = {
    "maiz": {"maize", "corn", "maiz"},
    "arroz": {"rice", "arroz"},
    "trigo": {"wheat", "trigo"},
    "soya": {"soybean", "soy", "soya", "soja"},
}

PHASE_LABELS = {1: "F1", 2: "F2", 3: "F3"}
PHASE_NAMES = {
    1: "F1 establecimiento / siembra",
    2: "F2 crecimiento vegetativo y reproductivo",
    3: "F3 maduracion y cosecha",
}
PHASE_COLORS = {1: "A7C957", 2: "7FB069", 3: "C08497"}
REQUIRED_PHASES = {1, 2, 3}

NON_PRODUCTIVE = [
    "out of season", "outofseason", "end season", "end of season", "endofseason",
    "harvest end", "off season", "no season", "fuera de temporada",
    "fuera temporada", "fuera_de_temporada", "fin de temporada", "fin_de_temporada",
]


def clean_text(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value):
    text = clean_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def detect_column(columns, meaning, candidates):
    normalized = {norm(c): c for c in columns}
    for candidate in candidates:
        if norm(candidate) in normalized:
            return normalized[norm(candidate)]
    tokens = [norm(c).split() for c in candidates]
    for col in columns:
        n = norm(col)
        if any(all(tok in n for tok in token_list) for token_list in tokens):
            return col
    return ""


def build_dictionary(df):
    specs = [
        ("cultivo", ["cultivo", "cultivo_normalizado", "crop_original", "crop"]),
        ("temporada", ["temporada", "season"]),
        ("region", ["region"]),
        ("pais", ["pais", "country"]),
        ("latitud", ["lat", "latitude", "latitud"]),
        ("franja latitudinal", ["banda_latitud", "lat_band", "franja"]),
        ("fase fenologica", ["fase_original", "phase_name", "fase_estandar"]),
        ("inicio de fase", ["phase_start_doy", "start_doy", "planting"]),
        ("fin de fase", ["phase_end_doy", "end_doy", "harvest"]),
        ("duracion", ["phase_duration", "duration", "duration_total"]),
        ("orden fase", ["phase_order"]),
        ("variable critica", ["variable_critica", "variable climatica critica"]),
        ("referencia bibliografica", ["fuente", "referencia", "enlace"]),
    ]
    rows = []
    columns = list(df.columns)
    for meaning, candidates in specs:
        col = detect_column(columns, meaning, candidates)
        rows.append({
            "nombre_original": col or "-",
            "significado_inferido": meaning,
            "uso_dentro_del_procesamiento": "usada" if col else "no encontrada / no usada",
        })
    return pd.DataFrame(rows)


def col_for(dictionary_df, meaning):
    row = dictionary_df[dictionary_df["significado_inferido"].eq(meaning)]
    if row.empty:
        return ""
    value = clean_text(row.iloc[0]["nombre_original"])
    return "" if value == "-" else value


def is_non_productive(value):
    n = norm(value)
    return any(term == n or term in n for term in [norm(x) for x in NON_PRODUCTIVE])


def crop_target(value):
    n = norm(value)
    for target, variants in CROP_TARGETS.items():
        if n in variants:
            return target
    return ""


def intervals_for_doy(start, end, wraps):
    if start > end:
        return [(start, 366), (1, end)]
    return [(start, end)]


def overlap_days(start, end, wraps, month_start, month_end):
    total = 0
    for a, b in intervals_for_doy(start, end, wraps):
        total += max(0, min(b, month_end) - max(a, month_start) + 1)
    return total


def dominant_phase(phase_rows, month_start, month_end):
    candidates = []
    for row in phase_rows:
        days = overlap_days(row["start"], row["end"], row["wraps"], month_start, month_end)
        if days:
            candidates.append((days, -row["order"], row["order"]))
    if not candidates:
        return 0
    candidates.sort(reverse=True)
    return candidates[0][2]


def logical_sequence(phase_rows):
    starts = {}
    ends = {}
    durations = {}
    for order in REQUIRED_PHASES:
        rows = [r for r in phase_rows if r["order"] == order]
        if not rows:
            return False, "fase faltante", starts, ends, durations
        starts[order] = int(round(pd.Series([r["start"] for r in rows]).median()))
        ends[order] = int(round(pd.Series([r["end"] for r in rows]).median()))
        durations[order] = int(round(pd.Series([r["duration"] for r in rows]).median()))
    anchor = starts[1]
    adjusted = {1: anchor}
    for order in [2, 3]:
        adjusted[order] = starts[order] if starts[order] >= anchor else starts[order] + 366
    if not (adjusted[1] <= adjusted[2] <= adjusted[3]):
        return False, "orden invertido por DOY", starts, ends, durations
    return True, "", starts, ends, durations


def signature_is_reasonable(signature):
    active = [x for x in signature if x]
    if sorted(set(active)) != [1, 2, 3]:
        return False, "fases mensuales incompletas"
    counts = Counter(active)
    if max(counts.values()) >= 6:
        return False, "fase domina seis meses o mas"
    return True, ""


def build_region_calendars(df, dictionary_df):
    crop_col = col_for(dictionary_df, "cultivo")
    season_col = col_for(dictionary_df, "temporada")
    country_col = col_for(dictionary_df, "pais")
    region_col = col_for(dictionary_df, "region")
    lat_col = col_for(dictionary_df, "latitud")
    phase_col = col_for(dictionary_df, "fase fenologica")
    start_col = col_for(dictionary_df, "inicio de fase")
    end_col = col_for(dictionary_df, "fin de fase")
    duration_col = col_for(dictionary_df, "duracion")
    order_col = col_for(dictionary_df, "orden fase")

    required_cols = [crop_col, season_col, country_col, region_col, lat_col, phase_col, start_col, end_col, order_col]
    if any(not c for c in required_cols):
        raise ValueError("No se pudieron identificar todas las columnas necesarias.")

    work = df.copy()
    work["cultivo_objetivo"] = work[crop_col].map(crop_target)
    work = work[work["cultivo_objetivo"].ne("")]
    work = work[~work[phase_col].map(is_non_productive)]
    work["phase_order_num"] = pd.to_numeric(work[order_col], errors="coerce")
    work["lat_num"] = pd.to_numeric(work[lat_col], errors="coerce")
    work["start_num"] = pd.to_numeric(work[start_col], errors="coerce")
    work["end_num"] = pd.to_numeric(work[end_col], errors="coerce")
    if duration_col:
        work["duration_num"] = pd.to_numeric(work[duration_col], errors="coerce")
    else:
        work["duration_num"] = None
    work = work.dropna(subset=["phase_order_num", "lat_num", "start_num", "end_num"])
    work["phase_order_num"] = work["phase_order_num"].astype(int)
    work = work[work["phase_order_num"].isin(REQUIRED_PHASES)]

    rows = []
    qc = []
    group_cols = ["cultivo_objetivo", season_col, country_col, region_col, "lat_num"]
    for key, group in work.groupby(group_cols, dropna=False):
        crop, season, country, region, lat = key
        phase_rows = []
        for _, r in group.iterrows():
            start = int(r["start_num"])
            end = int(r["end_num"])
            wraps = start > end
            duration = r["duration_num"]
            if pd.isna(duration):
                duration = sum(b - a + 1 for a, b in intervals_for_doy(start, end, wraps))
            phase_rows.append({
                "order": int(r["phase_order_num"]),
                "start": start,
                "end": end,
                "wraps": wraps,
                "duration": float(duration),
            })

        ok, reason, starts, ends, durations = logical_sequence(phase_rows)
        signature = []
        for _, mstart, mend in MONTHS:
            signature.append(dominant_phase(phase_rows, mstart, mend))
        reasonable, reason2 = signature_is_reasonable(signature)
        accepted = ok and reasonable
        qc.append({
            "cultivo": crop, "temporada": clean_text(season), "pais": clean_text(country),
            "region": clean_text(region), "lat": float(lat),
            "estado": "valido" if accepted else "excluido",
            "motivo": reason or reason2,
            "firma_mensual": "|".join(PHASE_LABELS.get(x, "") for x in signature),
        })
        if not accepted:
            continue

        record = {
            "cultivo": crop,
            "temporada": clean_text(season),
            "pais": clean_text(country),
            "region": clean_text(region),
            "lat": float(lat),
            "firma": "|".join(str(x) if x else "" for x in signature),
            "inicio_f1": starts[1], "inicio_f2": starts[2], "inicio_f3": starts[3],
            "fin_f1": ends[1], "fin_f2": ends[2], "fin_f3": ends[3],
            "dur_f1": durations[1], "dur_f2": durations[2], "dur_f3": durations[3],
        }
        for month, phase in zip(MONTH_NAMES, signature):
            record[month] = PHASE_LABELS.get(phase, "")
        rows.append(record)
    return pd.DataFrame(rows), pd.DataFrame(qc)


def similar(row_a, row_b):
    sig_a = [int(x) if x else 0 for x in clean_text(row_a["firma"]).split("|")]
    sig_b = [int(x) if x else 0 for x in clean_text(row_b["firma"]).split("|")]
    hamming = sum(a != b for a, b in zip(sig_a, sig_b))
    start_close = abs(float(row_a["inicio_f1"]) - float(row_b["inicio_f1"])) <= 15
    end_close = abs(float(row_a["fin_f3"]) - float(row_b["fin_f3"])) <= 15
    duration_close = all(abs(float(row_a[f"dur_f{i}"]) - float(row_b[f"dur_f{i}"])) <= 30 for i in [1, 2, 3])
    return hamming <= 1 or (start_close and end_close and duration_close)


def modal_signature(df):
    return df["firma"].value_counts().idxmax()


def build_typical_groups(region_df, min_rows=4, max_rows=8):
    output = []
    summary = []
    for (crop, season), group in region_df.groupby(["cultivo", "temporada"], dropna=False):
        pattern_groups = []
        for signature, cdf in group.groupby("firma", dropna=False):
            pattern_groups.append((len(cdf), float(cdf["lat"].max()), signature, cdf.copy()))
        pattern_groups = sorted(pattern_groups, key=lambda x: (-x[0], -x[1], x[2]))[:max_rows]

        for _, _, signature, cdf in pattern_groups:
            rep = cdf.iloc[0]
            match_pct = 100.0
            n = len(cdf)
            row = {
                "cultivo": crop,
                "temporada": clean_text(season),
                "franja_latitudinal": f"{round(float(cdf['lat'].max()), 1)} a {round(float(cdf['lat'].min()), 1)}",
                "lat_min": round(float(cdf["lat"].min()), 2),
                "lat_max": round(float(cdf["lat"].max()), 2),
                "n_regiones": n,
                "cobertura_pct": round(n / len(group) * 100, 1) if len(group) else 0,
                "coincidencia_patron_pct": match_pct,
                "duracion_prom_f1": round(float(cdf["dur_f1"].mean()), 1),
                "duracion_prom_f2": round(float(cdf["dur_f2"].mean()), 1),
                "duracion_prom_f3": round(float(cdf["dur_f3"].mean()), 1),
                "desv_duracion_f1": round(float(cdf["dur_f1"].std(ddof=0)), 1),
                "desv_duracion_f2": round(float(cdf["dur_f2"].std(ddof=0)), 1),
                "desv_duracion_f3": round(float(cdf["dur_f3"].std(ddof=0)), 1),
                "firma": signature,
                "paises_ejemplo": ", ".join(sorted(set(cdf["pais"].map(clean_text)))[:8]),
            }
            for month in MONTH_NAMES:
                row[month] = clean_text(rep[month])
            output.append(row)
            summary.append({k: row[k] for k in [
                "cultivo", "temporada", "franja_latitudinal", "lat_min", "lat_max",
                "n_regiones", "cobertura_pct", "coincidencia_patron_pct",
                "duracion_prom_f1", "duracion_prom_f2", "duracion_prom_f3",
                "desv_duracion_f1", "desv_duracion_f2", "desv_duracion_f3",
                "paises_ejemplo",
            ]})
    return pd.DataFrame(output), pd.DataFrame(summary)


def write_df_sheet(ws, df, title=None):
    start = 1
    if title:
        ws.cell(1, 1, title).font = Font(bold=True, size=14)
        start = 3
    for col_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(start, col_idx, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="334155")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_idx, (_, row) in enumerate(df.iterrows(), start + 1):
        for col_idx, col in enumerate(df.columns, 1):
            ws.cell(row_idx, col_idx, clean_text(row[col]))
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max(1, ws.max_column)):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(ws.cell(start, col_idx).value)) + 2, 12), 28)


def write_calendar_sheet(ws, crop_df, crop_label):
    ws.cell(1, 1, f"Calendario tipico - {crop_label}").font = Font(bold=True, size=14)
    for idx, order in enumerate([1, 2, 3], 3):
        cell = ws.cell(2, idx, PHASE_NAMES[order])
        cell.fill = PatternFill("solid", fgColor=PHASE_COLORS[order])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    row_cursor = 4
    thin = Side(style="thin", color="D9D9D9")
    for season, group in crop_df.groupby("temporada", dropna=False):
        ws.cell(row_cursor, 1, f"Temporada {season}").font = Font(bold=True, size=12)
        row_cursor += 1
        headers = ["Franja latitudinal", "n", "cobertura %", "coincidencia %", *MONTH_NAMES]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row_cursor, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        row_cursor += 1
        for _, row in group.iterrows():
            ws.cell(row_cursor, 1, clean_text(row["franja_latitudinal"]))
            ws.cell(row_cursor, 2, clean_text(row["n_regiones"]))
            ws.cell(row_cursor, 3, clean_text(row["cobertura_pct"]))
            ws.cell(row_cursor, 4, clean_text(row["coincidencia_patron_pct"]))
            ws.cell(row_cursor, 2).comment = Comment(clean_text(row["paises_ejemplo"]), "Codex")
            for idx, month in enumerate(MONTH_NAMES, 5):
                label = clean_text(row[month])
                cell = ws.cell(row_cursor, idx, label)
                if label:
                    order = int(label.replace("F", ""))
                    cell.fill = PatternFill("solid", fgColor=PHASE_COLORS[order])
                cell.alignment = Alignment(horizontal="center")
            row_cursor += 1
        row_cursor += 2
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=16):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    widths = {1: 18, 2: 8, 3: 12, 4: 14}
    for col_idx in range(1, 17):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 11)
    ws.freeze_panes = "E6"


def main():
    df = pd.read_excel(INPUT, sheet_name="tabla_maestra", dtype=object)
    dictionary_df = build_dictionary(df)
    region_df, qc_df = build_region_calendars(df, dictionary_df)
    typical_df, summary_df = build_typical_groups(region_df)

    wb = Workbook()
    wb.remove(wb.active)
    sheet_names = {
        "maiz": "calendario_tipico_maiz",
        "arroz": "calendario_tipico_arroz",
        "trigo": "calendario_tipico_trigo",
        "soya": "calendario_tipico_soya",
    }
    for crop, sheet_name in sheet_names.items():
        ws = wb.create_sheet(sheet_name)
        crop_df = typical_df[typical_df["cultivo"].eq(crop)].copy()
        if crop_df.empty:
            ws.cell(1, 1, f"No hay calendario fenologico completo disponible para {crop}.").font = Font(bold=True)
        else:
            write_calendar_sheet(ws, crop_df, crop)

    write_df_sheet(wb.create_sheet("resumen_agrupaciones"), summary_df, "Resumen de agrupaciones")
    write_df_sheet(wb.create_sheet("control_calidad"), qc_df, "Control de calidad regional")
    write_df_sheet(wb.create_sheet("diccionario_columnas"), dictionary_df, "Diccionario de columnas")

    output = OUT / "calendarios_fenologicos_tipicos_web.xlsx"
    wb.save(output)
    typical_df.to_csv(OUT / "calendarios_fenologicos_tipicos_web.csv", index=False, encoding="utf-8-sig")
    summary_df.to_csv(OUT / "resumen_agrupaciones_calendarios.csv", index=False, encoding="utf-8-sig")
    qc_df.to_csv(OUT / "control_calidad_calendarios.csv", index=False, encoding="utf-8-sig")
    dictionary_df.to_csv(OUT / "diccionario_columnas_calendarios.csv", index=False, encoding="utf-8-sig")
    print(output)
    print(typical_df.groupby(["cultivo", "temporada"]).size().to_string())
    print(qc_df["estado"].value_counts().to_string())


if __name__ == "__main__":
    main()
