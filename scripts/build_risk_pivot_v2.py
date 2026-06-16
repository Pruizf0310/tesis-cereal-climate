from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


DEFAULT_INPUT = Path(r"C:\Users\paola\Tesis\03_Resultados\Fenologia\geoglam_cm4ew_tabla_maestra.xlsx")
DEFAULT_EXCEL_OUTPUT = Path("outputs/geoglam_cm4ew_tabla_maestra_risk_v2.xlsx")
DEFAULT_JSON_OUTPUT = Path("web-v2/public/data/risk_pivot_v2.json")

BASE_COLUMNS = [
    "cultivo",
    "fase_original",
    "fase_estandar",
    "phase_name",
    "variable_critica",
    "umbral",
    "tipo_estres",
    "impacto_cultivo",
    "nivel_evidencia",
    "fuente",
    "enlace",
    "matched_sensib_clima",
    "matched_fenologia",
]

OUTPUT_COLUMNS = [
    "cultivo",
    "etapa_derivada",
    "amenaza",
    "umbral",
    "impacto_cualitativo",
    "impacto_cuantitativo",
    "categoria_impacto",
    "evidencia",
    "fuente",
    "enlace",
    "criterio_calculo",
    "registros_base",
]

WEB_STAGE_LABELS = {
    "Floración / reproducción": "Flowering / reproduction",
    "Germinación / establecimiento": "Germination / establishment",
    "Llenado de grano / formación de rendimiento": "Grain filling / yield formation",
    "Llenado de grano / maduración": "Grain filling / maturation",
    "Maduración / cosecha": "Maturation / harvest",
    "Desarrollo vegetativo": "Vegetative development",
    "No determinado": "Not determined",
}

WEB_HAZARD_LABELS = {
    "Déficit hídrico en ventana reproductiva — Hídrico": "Water deficit during the reproductive window — Water stress",
    "Humedad del suelo en zona radicular — Hídrico": "Root-zone soil moisture — Water stress",
    "Temperatura durante llenado efectivo del grano — Térmico": "Temperature during effective grain filling — Heat stress",
    "Temperatura máxima durante antesis/floración — Térmico": "Maximum temperature during anthesis/flowering — Heat stress",
    "Potencial hídrico del suelo/solución — Hídrico": "Soil/solution water potential — Water stress",
    "Temperatura del aire durante llenado del grano — Térmico": "Air temperature during grain filling — Heat stress",
    "Temperatura del aire en antesis — Térmico": "Air temperature at anthesis — Heat stress",
    "Temperatura durante emergencia y crecimiento otoñal temprano — Térmico": "Temperature during emergence and early autumn growth — Heat stress",
    "Temperatura del aire en llenado medio a terminal — Térmico": "Air temperature during mid-to-terminal grain filling — Heat stress",
}

WEB_THRESHOLD_LABELS = {
    "<30 mm en 10 días durante fase reproductiva en maíz grano (umbral empírico distrital); evidencia experimental complementaria con ψs ≈ −50 kPa como estrés manejado":
        "<30 mm over 10 days during the reproductive phase in grain maize (district-scale empirical threshold); complementary experimental evidence with ψs ≈ −50 kPa as managed stress",
    "θ < 0.183 cm³ cm⁻³ en 0–1 m, derivado con criterio FAO (p = 0.55) a partir de θFC = 0.26 y θWP = 0.12; por debajo de ese valor comenzó (K_s<1)":
        "θ < 0.183 cm³ cm⁻³ in 0-1 m soil depth, derived with the FAO criterion (p = 0.55) from θFC = 0.26 and θWP = 0.12; below this value, stress began (K_s<1)",
    "Tratamientos con Tmax media de 39.4–41.5 °C durante 7 días en llenado efectivo generaron daño claro; a 32.9 °C no hubo efecto significativo":
        "Treatments with mean Tmax of 39.4-41.5 °C for 7 days during effective grain filling caused clear damage; at 32.9 °C there was no significant effect",
    "≥35 °C durante antesis; en el experimento, el tratamiento térmico tuvo pico de 39 °C y se aplicó 48 h":
        "≥35 °C during anthesis; in the experiment, the heat treatment peaked at 39 °C and was applied for 48 h",
    "−0.046 a −0.056 MPa: umbral a partir del cual comienzan a caer evapotranspiración, expansión foliar y biomasa; en campo, la productividad cayó cuando se alcanzó −0.05 a −0.06 MPa":
        "−0.046 to −0.056 MPa: threshold where evapotranspiration, leaf expansion and biomass begin to decline; in field conditions, productivity declined at −0.05 to −0.06 MPa",
    "T media diaria >25 °C durante llenado se asocia a pérdida de calidad; en campo, un aumento de 1.6–3.1 °C durante llenado redujo el rendimiento":
        "Mean daily temperature >25 °C during filling is associated with quality loss; in the field, a 1.6-3.1 °C increase during filling reduced yield",
    "35/25 °C día/noche durante 7 días en antesis": "35/25 °C day/night for 7 days at anthesis",
    "<10 °C o >30 °C reducen significativamente la germinación; alrededor de 8 °C aún germina, pero con emergencia muy lenta y menor rendimiento posterior":
        "<10 °C or >30 °C significantly reduce germination; around 8 °C germination still occurs, but emergence is very slow and later yield is lower",
    "38/28 °C día/noche durante 7 días en llenado medio; además, >30 °C se asocia de forma consistente con caída del llenado":
        "38/28 °C day/night for 7 days during mid grain filling; >30 °C is also consistently associated with reduced filling",
}

WEB_IMPACT_LABELS = {
    "Bajo": "Low",
    "Moderado": "Moderate",
    "Alto": "High",
    "Crítico": "Critical",
    "No determinado": "Not determined",
}

WEB_EVIDENCE_LABELS = {
    "Modelación empírica a escala distrital + experimento de campo": "District-scale empirical modeling + field experiment",
    "Experimental de campo multianual + umbral operativo derivado por balance hídrico":
        "Multi-year field experiment + operational threshold derived from water balance",
    "Experimental de campo": "Field experiment",
    "Experimental en ambiente controlado": "Controlled-environment experiment",
    "Experimental en invernadero con validación de campo": "Greenhouse experiment with field validation",
    "Experimental de campo + revisión": "Field experiment + review",
    "Experimental de campo multianual + síntesis fisiológica": "Multi-year field experiment + physiological synthesis",
}

SEVERITY_ORDER = {"No determinado": 0, "Bajo": 1, "Moderado": 2, "Alto": 3, "Crítico": 4}
EVIDENCE_STRENGTH = {
    "Experimental de campo": 5,
    "Validación de campo": 5,
    "Experimental": 4,
    "Ambiente controlado": 4,
    "Invernadero": 4,
    "Revisión": 3,
    "Review": 3,
    "Literatura": 3,
}


def norm(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.casefold()


def display(value: Any, default: str = "No determinado") -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return default
    text = str(value).strip()
    return text if text else default


def is_true(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return norm(value) in {"true", "1", "si", "sí", "yes", "y"}


STAGE_RULES = [
    (
        "Germinación / establecimiento",
        ["germinacion", "emergencia", "plantula", "establecimiento", "semilla", "seedling"],
    ),
    (
        "Desarrollo vegetativo",
        [
            "vegetativo",
            "tillering",
            "macollamiento",
            "hojas",
            "area foliar",
            "expansion foliar",
            "crecimiento vegetativo",
            "biomasa temprana",
        ],
    ),
    (
        "Floración / reproducción",
        [
            "floracion",
            "antesis",
            "polinizacion",
            "espigamiento",
            "panicula",
            "panoja",
            "silk",
            "tassel",
            "reproductive",
            "seed-setting",
            "cuajado",
            "esterilidad",
        ],
    ),
    (
        "Llenado de grano / formación de rendimiento",
        [
            "grano",
            "llenado",
            "grain filling",
            "grain weight",
            "peso de grano",
            "semilla",
            "pod filling",
        ],
    ),
    (
        "Maduración / cosecha",
        ["maduracion", "cosecha", "senescencia", "harvest", "calidad de grano", "humedad en cosecha"],
    ),
]

MACRO_STAGE = {
    "siembra_vegetativa_temprana": "Germinación / establecimiento",
    "siembra a vegetativa temprana": "Germinación / establecimiento",
    "vegetativa_reproductiva": "Floración / reproducción",
    "vegetativa a reproductiva": "Floración / reproducción",
    "maduracion_cosecha": "Llenado de grano / maduración",
    "maduracion a cosecha": "Llenado de grano / maduración",
}


def derive_stage(row: pd.Series) -> tuple[str, str]:
    high_text = norm(f"{row.get('impacto_cultivo', '')} {row.get('variable_critica', '')}")
    all_text = norm(
        " ".join(
            display(row.get(col), "")
            for col in [
                "fase_original",
                "fase_estandar",
                "phase_name",
                "variable_critica",
                "tipo_estres",
                "impacto_cultivo",
            ]
        )
    )
    for stage, keywords in STAGE_RULES:
        if any(keyword in high_text for keyword in keywords):
            return stage, "Alta"
    for stage, keywords in STAGE_RULES:
        if any(keyword in all_text for keyword in keywords):
            return stage, "Baja"
    for key, stage in MACRO_STAGE.items():
        if key in norm(row.get("fase_estandar")) or key in norm(row.get("fase_original")) or key in norm(row.get("phase_name")):
            return stage, "Media"
    return "No determinado", "Baja"


def build_hazard(row: pd.Series) -> str:
    variable = display(row.get("variable_critica"))
    stress = display(row.get("tipo_estres"))
    if stress == "No determinado":
        return variable
    return f"{variable} — {stress}"


WEIGHT_KEYWORDS = [
    (1.0, ["rendimiento", "yield", "productividad", "production"]),
    (
        0.9,
        [
            "cuajado",
            "seed-setting",
            "esterilidad",
            "polinizacion",
            "grain filling",
            "llenado",
            "peso de grano",
            "grain weight",
        ],
    ),
    (0.7, ["biomasa", "biomass", "root biomass", "shoot biomass"]),
    (0.6, ["area foliar", "expansion foliar", "leaf area", "crecimiento vegetativo"]),
    (0.5, ["fotosintesis", "photosynthesis", "conductancia", "stomatal conductance"]),
    (0.4, ["evapotranspiracion", " et", "eficiencia de uso del agua", "water use efficiency", "wue"]),
]


def impact_weight(text: str) -> float:
    normalized = norm(text)
    for weight, keywords in WEIGHT_KEYWORDS:
        if any(keyword in normalized for keyword in keywords):
            return weight
    return 0.3


def evidence_factor(evidence: Any) -> float:
    text = norm(evidence)
    if "campo" in text or "validacion de campo" in text:
        return 1.2
    if any(keyword in text for keyword in ["experimental", "ambiente controlado", "invernadero"]):
        return 1.0
    if any(keyword in text for keyword in ["revision", "review", "literatura"]):
        return 0.9
    return 0.8


PERCENT_RE = re.compile(
    r"(?<![\w])[-−–]?\s*(\d+(?:[.,]\d+)?)\s*(?:%|(?:[–—-]| a | to )\s*[-−–]?\s*(\d+(?:[.,]\d+)?)\s*%)",
    flags=re.IGNORECASE,
)


def number(value: str) -> float:
    return abs(float(value.replace(",", ".")))


def context_window(text: str, start: int, end: int, radius: int = 90) -> str:
    return text[max(0, start - radius) : min(len(text), end + radius)]


def quantitative_impact(row: pd.Series) -> tuple[float | None, str]:
    text = display(row.get("impacto_cultivo"), "")
    matches = list(PERCENT_RE.finditer(text))
    if not matches:
        return None, "Regla cualitativa: no se encontró porcentaje explícito en impacto_cultivo."
    factor = evidence_factor(row.get("nivel_evidencia"))
    candidates = []
    for match in matches:
        values = [number(match.group(1))]
        if match.group(2):
            values.append(number(match.group(2)))
        pct = max(values)
        weight = impact_weight(context_window(text, match.start(), match.end()))
        candidates.append(min(100.0, pct * weight * factor))
    return max(candidates), "Porcentaje extraído de impacto_cultivo, ponderado por tipo de impacto y evidencia."


def qualitative_impact(text_value: Any, quant: float | None) -> str:
    text = norm(text_value)
    if any(
        keyword in text
        for keyword in [
            "muerte",
            "perdida severa",
            "esterilidad",
            "falla reproductiva",
            "aborto floral",
            "fallo de polinizacion",
            "colapso de llenado",
            "no recuperable",
        ]
    ):
        return "Crítico"
    if quant is not None:
        if quant > 50:
            return "Crítico"
        if quant > 30:
            return "Alto"
        if quant > 15:
            return "Moderado"
        return "Bajo"
    if any(keyword in text for keyword in ["rendimiento", "yield", "biomasa", "area foliar", "cuajado", "llenado", "productividad", "calidad"]):
        return "Alto"
    if any(keyword in text for keyword in ["fotosintesis", "eficiencia de uso del agua", "evapotranspiracion", "crecimiento", "conductancia"]):
        return "Moderado"
    if any(keyword in text for keyword in ["preventivo", "suboptimo", "leve"]):
        return "Bajo"
    return "No determinado"


def quantitative_category(value: float | None) -> str:
    if value is None:
        return "No determinado"
    if value <= 15:
        return "Bajo"
    if value <= 30:
        return "Moderado"
    if value <= 50:
        return "Alto"
    return "Crítico"


def strongest_evidence(values: pd.Series) -> str:
    unique = unique_join(values)
    best_value = "No determinado"
    best_score = -1
    for value in values.dropna().astype(str):
        value_score = 1
        normalized = norm(value)
        for keyword, score in EVIDENCE_STRENGTH.items():
            if norm(keyword) in normalized:
                value_score = max(value_score, score)
        if value_score > best_score:
            best_score = value_score
            best_value = value
    return best_value if best_value != "No determinado" else unique


def unique_join(values: pd.Series) -> str:
    seen: list[str] = []
    for value in values:
        text = display(value, "")
        if text and text not in seen:
            seen.append(text)
    return "; ".join(seen) if seen else "No determinado"


def summarize_criteria(values: pd.Series) -> str:
    criteria = set(display(value, "") for value in values)
    has_quant = any("Porcentaje extraído" in value for value in criteria)
    has_qual = any("Regla cualitativa" in value for value in criteria)
    if has_quant and has_qual:
        return "Cálculo mixto: porcentaje extraído cuando existió y regla cualitativa para registros sin porcentaje explícito."
    if has_quant:
        return "Porcentaje extraído de impacto_cultivo, ponderado por tipo de impacto y evidencia."
    return "Regla cualitativa: no se encontró porcentaje explícito en impacto_cultivo."


def build_pivot(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    missing = [column for column in BASE_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(missing)}")
    if "record_id" not in df.columns:
        df = df.copy()
        df["record_id"] = range(1, len(df) + 1)

    base = df[
        df["matched_sensib_clima"].map(is_true)
        & df["matched_fenologia"].map(is_true)
        & df["variable_critica"].map(lambda value: display(value, "") != "")
        & df["umbral"].map(lambda value: display(value, "") != "")
        & df["impacto_cultivo"].map(lambda value: display(value, "") != "")
    ].copy()

    stages = base.apply(derive_stage, axis=1)
    base["etapa_derivada"] = [stage for stage, _confidence in stages]
    base["confianza_etapa"] = [confidence for _stage, confidence in stages]
    base["amenaza"] = base.apply(build_hazard, axis=1)
    impacts = base.apply(quantitative_impact, axis=1)
    base["impacto_cuantitativo_raw"] = [value for value, _criterion in impacts]
    base["criterio_calculo"] = [criterion for _value, criterion in impacts]
    base["impacto_cualitativo"] = base.apply(
        lambda row: qualitative_impact(row["impacto_cultivo"], row["impacto_cuantitativo_raw"]),
        axis=1,
    )

    grouped_rows: list[dict[str, Any]] = []
    for (crop, stage, hazard, threshold), group in base.groupby(
        ["cultivo", "etapa_derivada", "amenaza", "umbral"], dropna=False, sort=True
    ):
        quant_values = group["impacto_cuantitativo_raw"].dropna()
        max_quant = None if quant_values.empty else float(quant_values.max())
        qualitative = max(group["impacto_cualitativo"], key=lambda value: SEVERITY_ORDER.get(value, 0))
        grouped_rows.append(
            {
                "cultivo": crop,
                "etapa_derivada": stage,
                "amenaza": hazard,
                "umbral": threshold,
                "impacto_cualitativo": qualitative,
                "impacto_cuantitativo": "No determinado" if max_quant is None else round(max_quant, 2),
                "categoria_impacto": quantitative_category(max_quant),
                "evidencia": strongest_evidence(group["nivel_evidencia"]),
                "fuente": unique_join(group["fuente"]),
                "enlace": unique_join(group["enlace"]),
                "criterio_calculo": summarize_criteria(group["criterio_calculo"]),
                "registros_base": "; ".join(str(value) for value in sorted(set(group["record_id"].astype(str)))),
            }
        )

    pivot = pd.DataFrame(grouped_rows, columns=OUTPUT_COLUMNS).sort_values(
        ["cultivo", "etapa_derivada", "amenaza", "umbral"]
    )
    summary = {
        "cultivos_procesados": base["cultivo"].nunique(),
        "registros_originales_usados": len(base),
        "filas_finales": len(pivot),
        "impactos_cuantitativos_calculados": int(pivot["impacto_cuantitativo"].ne("No determinado").sum()),
        "impactos_no_determinados": int(pivot["impacto_cuantitativo"].eq("No determinado").sum()),
    }
    return pivot, summary


def write_excel(input_path: Path, output_path: Path, pivot: pd.DataFrame) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)

    workbook = load_workbook(output_path)
    if "risk_pivot_v2" in workbook.sheetnames:
        del workbook["risk_pivot_v2"]
    sheet = workbook.create_sheet("risk_pivot_v2")
    sheet.append(OUTPUT_COLUMNS)
    for row in pivot.itertuples(index=False):
        sheet.append(list(row))
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 70)
    workbook.save(output_path)


def web_criteria(value: str) -> str:
    if value.startswith("Porcentaje extraído"):
        return "Percentage extracted from impact_cultivo, weighted by impact type and evidence level."
    if value.startswith("Regla cualitativa"):
        return "Qualitative rule: no explicit percentage was found in impact_cultivo."
    if value.startswith("Cálculo mixto"):
        return "Mixed calculation: extracted percentage where available and qualitative rule for records without an explicit percentage."
    return value


def web_record(record: dict[str, Any]) -> dict[str, Any]:
    converted = dict(record)
    converted["etapa_derivada"] = WEB_STAGE_LABELS.get(record["etapa_derivada"], record["etapa_derivada"])
    converted["amenaza"] = WEB_HAZARD_LABELS.get(record["amenaza"], record["amenaza"])
    converted["umbral"] = WEB_THRESHOLD_LABELS.get(record["umbral"], record["umbral"])
    converted["impacto_cualitativo"] = WEB_IMPACT_LABELS.get(record["impacto_cualitativo"], record["impacto_cualitativo"])
    converted["categoria_impacto"] = WEB_IMPACT_LABELS.get(record["categoria_impacto"], record["categoria_impacto"])
    if converted["impacto_cuantitativo"] == "No determinado":
        converted["impacto_cuantitativo"] = "Not determined"
    converted["evidencia"] = WEB_EVIDENCE_LABELS.get(record["evidencia"], record["evidencia"])
    converted["criterio_calculo"] = web_criteria(record["criterio_calculo"])
    return converted


def write_json(output_path: Path, pivot: pd.DataFrame) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = [web_record(record) for record in pivot.to_dict(orient="records")]
    output_path.write_text(json.dumps(records, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build risk_pivot_v2 sheet and JSON for the /risk page.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Source Excel workbook.")
    parser.add_argument("--excel-output", type=Path, default=DEFAULT_EXCEL_OUTPUT, help="Workbook to update or create.")
    parser.add_argument("--json-output", type=Path, default=DEFAULT_JSON_OUTPUT, help="JSON data file for web-v2.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input
    excel_output = args.excel_output
    json_output = args.json_output

    df = pd.read_excel(input_path, sheet_name="tabla_maestra")
    pivot, summary = build_pivot(df)
    write_excel(input_path, excel_output, pivot)
    write_json(json_output, pivot)

    print("risk_pivot_v2 generado")
    print(f"Cultivos procesados: {summary['cultivos_procesados']}")
    print(f"Registros originales usados: {summary['registros_originales_usados']}")
    print(f"Filas finales en risk_pivot_v2: {summary['filas_finales']}")
    print(f"Impactos cuantitativos calculados: {summary['impactos_cuantitativos_calculados']}")
    print(f"Impactos no determinados: {summary['impactos_no_determinados']}")
    print(f"Excel actualizado: {excel_output}")
    print(f"JSON exportado: {json_output}")


if __name__ == "__main__":
    main()
